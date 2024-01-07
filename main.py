from openpyxl import load_workbook
wb = load_workbook("friends.xlsx")
ws = wb.active

rows = []
headers = []

for i in range(1, ws.max_column+1):
    cell = ws.cell(row=1, column=i)
    headers.append(cell.value)

for i in range(2, ws.max_row+1):
    row = [cell.value for cell in ws[i]]
    rows.append(row)

table_name = wb.sheetnames[0]

SQL = f"CREATE TABLE {table_name} ("

for header in headers:
	SQL = SQL + "`" + header + "`" +" VARCHAR(128),"
 
SQL = SQL[:-1] + ");"
print("\n")

SQL2 = f"INSERT INTO {table_name} ("

for header in headers:
	SQL2 = SQL2 + "`" + header + "`,"
SQL2 = SQL2[:-1]


SQL2 = SQL2 + ") VALUES "



for row in rows:
	SQL2 = SQL2 + "("
	for el in row:
		SQL2 = SQL2 + "'" + str(el) + "',"
	SQL2 = SQL2[:-1]
	SQL2 = SQL2 + "),"
SQL2 = SQL2[:-1]


SQL2 = SQL2 + ";"


print(SQL)

print(SQL2)
